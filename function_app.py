import logging
import internal
import datetime
import azure.functions as func

from openpyxl import load_workbook
from decimal import Decimal, getcontext


app = func.FunctionApp()

getcontext().prec = 4


@app.schedule(schedule="1 * * * * *", arg_name="my_timer", run_on_startup=True,
              use_monitor=False) 
def zmb_52_script(my_timer: func.TimerRequest) -> None:
    if my_timer.past_due:
        logging.info('The timer is past due!')

    logging.info('Python timer trigger function executed.')


def zmb_52_script_local():
    # Load the workbook
    wb = load_workbook('ZMB52 Nacional_2023_09_Nov_v2.xlsx')

    # Get the sheet
    sheet = wb.active

    current_execute_date_less_one = datetime.datetime.now() - datetime.timedelta(days=1)

    # Read each row in the sheet
    for idx_row, row in enumerate(sheet.iter_rows(min_row=6, min_col=2)):

        country: str = ""
        part_suffix: str = ""
        part_number: str = ""
        material: str = ""
        part_description: str = ""
        base_unit_of_measure: str = ""
        total_quantity_stock: float = 0.0
        total_unrestricted_stock: float = 0.0
        sales_last_52_weeks: Decimal = Decimal(0.0)
        last_good_receipt = None
        lasted_sales_date = None
        marketing_code: str = ""
        product_category: str = ""
        product_code: str = ""
        shelf_life_code: int = 0
        material_status: str = ""
        part_day: int = 0
        part_age: Decimal = Decimal(0.0)
        total_stock_value: Decimal = Decimal(0.00)
        total_consignment_stock: Decimal = Decimal(0.000)
        total_blocked_stock: Decimal = Decimal(0.000)
        moving_average_price: Decimal = Decimal(0.00)
        net_purchase_price: Decimal = Decimal(0.00)
        lasted_delivery_date = None
        lasted_inventory_date = None
        total_stock_in_transit: Decimal = Decimal(0.000)
        report_date = current_execute_date_less_one
        snapshot_date = datetime.datetime.now()

        # Read each cell in the row
        for idx_cell, cell in enumerate(row):
            if idx_cell == 0:
                country = cell.value

            if idx_cell == 1:
                part_suffix = cell.value

            if idx_cell == 2:
                part_number = str(cell.value)

            if idx_cell == 3:
                material = str(part_number) + ":" + str(part_suffix)

            if idx_cell == 4:
                part_description = cell.value

            if idx_cell == 5:
                base_unit_of_measure = str(cell.value).replace("ST", "PC")

            # TODO ask to Rodrigo if this is correct
            #if idx_cell == 6:
            #    total_quantity_stock = float(cell.value)

            if idx_cell == 7:
                total_unrestricted_stock = float(cell.value)

            if idx_cell == 8:
                if "-" in str(cell.value):
                    t = cell.value
                    sales_last_52_weeks = Decimal(float(cell.value.replace("-", "")) * -1)
                else:
                    sales_last_52_weeks = Decimal(cell.value)

            if idx_cell == 9:
                if cell.value == 0:
                    last_good_receipt = None
                else:
                    last_good_receipt = cell.value

            if idx_cell == 10:
                if cell.value == 0:
                    lasted_sales_date = None
                else:
                    lasted_sales_date = cell.value

            if idx_cell == 11:
                marketing_code = cell.value

            if idx_cell == 12:
                product_category = cell.value

            if idx_cell == 13:
                product_code = cell.value

            if idx_cell == 14:
                shelf_life_code = cell.value

            if idx_cell == 15:
                material_status = cell.value

            if idx_cell == 16:
                value_dates = internal.return_day_calculate_diff_day_from_days(last_good_receipt)
                part_day = value_dates

            if idx_cell == 17:
                getcontext().prec = 4
                value_dates = internal.return_age_calculate_diff_day_from_days(last_good_receipt)
                if value_dates is not None:
                    part_age_converted = Decimal(value_dates)
                    part_age = part_age_converted
                else:
                    part_age = None

            if idx_cell == 18:
                total_stock_value = Decimal(cell.value)

            if idx_cell == 19:
                if cell.value is None:
                    total_consignment_stock = Decimal(0.000)
                else:
                    total_consignment_stock = Decimal(cell.value)

            if idx_cell == 20:
                if cell.value is None:
                    total_blocked_stock = Decimal(0.000)
                else:
                    total_blocked_stock = Decimal(cell.value)

            if idx_cell == 21:
                if cell.value is None:
                    moving_average_price = Decimal(0.00)
                else:
                    moving_average_price = Decimal(cell.value)

            if idx_cell == 22:
                if cell.value is None:
                    net_purchase_price = Decimal(0.00)
                else:
                    net_purchase_price = Decimal(cell.value)

            if idx_cell == 23:
                if cell.value == 0:
                    lasted_delivery_date = None
                else:
                    lasted_delivery_date = cell.value

            if idx_cell == 24:
                if cell.value == 0:
                    lasted_inventory_date = None
                else:
                    lasted_inventory_date = cell.value

            if idx_cell == 25:
                if cell.value is None:
                    total_stock_in_transit = Decimal(0.000)
                else:
                    total_stock_in_transit = Decimal(cell.value)

        print(f"country: {country} "
              f"part_suffix: {part_suffix} "
              f"part_number: {part_number} "
              f"material: {material} "
              f"part_description: {part_description} "
              f"base_unit_of_measure: {base_unit_of_measure} "
              f"total_quantity_stock: {total_quantity_stock} "
              f"total_unrestricted_stock: {total_unrestricted_stock} "
              f"sales_last_52_weeks: {sales_last_52_weeks} "
              f"last_good_receipt: {last_good_receipt}")


zmb_52_script_local()
